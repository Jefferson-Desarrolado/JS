import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.cell import MergedCell
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill

# Datos proporcionados
data = {
    'FECHA': ['3/17/2025', None, '3/18/2025', '3/18/2025', None, None, '3/24/2025', None, '3/24/2025', None],
    'PACIENTE': ['NILO SERNAQUE', None, 'DIANA MACIAS', 'ALEXANDER ATIENCIA', None, None, 'DAVID CHIRIGUAYA', None, 'ALEXANDER ATIENCIA', None],
    'PROCEDIMIENTO': ['IMPLANTE CAPILAR COMPLETO', None, 'IMPLANTE DE CEJAS', 'IMPLANTE DE CEJAS', 'PORCENTAJE DE BARBA,', 'BIGOTE', 'IMPLANTE DE LINEA FRONTAL', None, 'IMPLANTE DE CEJAS (SESION EXTRA)', None],
    'PRECIO': ['$1,378', None, '$200', '$400', None, None, '$800', None, '$150', None],
    'HORARIO': ['9AM - 10PM', None, '9AM - 11PM', '11AM - 11PM', None, None, '9AM - 6PM', None, '6PM A 11PM', None]
}

# Crear DataFrame de pandas
df = pd.DataFrame(data)

# Generar Archivo Excel con Diseño Elaborado
def generar_excel_elaborado(data, filename="agenda_cosmetologia_elaborada.xlsx"):
    df = pd.DataFrame(data)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Giras Guayaquil - Cosmetología"

    # --- Definir Paleta de Colores Profesional ---
    color_primario = "4A148C"  # Deep Purple
    color_secundario = "E040FB" # Purple A400
    color_fondo = "F3E5F5"     # Lilac
    color_texto_header = "FFFFFF"
    color_borde = "9C27B0"      # Purple 500
    color_alterno = "FCE4EC"    # Pink 50

    # --- Estilos ---
    titulo_font = Font(size=26, bold=True, color=color_primario)
    titulo_alignment = Alignment(horizontal='center', vertical='center')
    subtitulo_font = Font(size=16, italic=True, color=color_primario)
    subtitulo_alignment = Alignment(horizontal='center', vertical='center')
    header_font = Font(bold=True, color=color_texto_header)
    header_alignment = Alignment(horizontal='center', vertical='center')
    header_fill = PatternFill(start_color=color_primario, end_color=color_primario, fill_type="solid")
    header_border = Border(bottom=Side(style='thick', color=color_secundario))
    data_font = Font(color=color_primario)
    data_alignment = Alignment(horizontal='center', vertical='center')
    data_border = Border(bottom=Side(style='thin', color=color_borde))
    alterno_fill = PatternFill(start_color=color_alterno, end_color=color_alterno, fill_type="solid")

    # --- Aplicar Estilos y Datos ---
    # Título
    sheet.merge_cells("A1:E1")
    title_cell = sheet['A1']
    title_cell.value = "GIRAS GUAYAQUIL"
    title_cell.font = titulo_font
    title_cell.alignment = titulo_alignment
    title_cell.fill = PatternFill(start_color=color_fondo, end_color=color_fondo, fill_type="solid")
    sheet.row_dimensions[1].height = 35

    # Subtítulo
    sheet.merge_cells("B2:C2")
    subtitle_cell = sheet['B2']
    subtitle_cell.value = "Agenda de Servicios de Cosmetología"
    subtitle_cell.font = subtitulo_font
    subtitle_cell.alignment = subtitulo_alignment
    sheet.row_dimensions[2].height = 25

    # Encabezados
    headers = ['FECHA', 'PACIENTE', 'PROCEDIMIENTO', 'PRECIO', 'HORARIO']
    for col_num, header in enumerate(headers, start=1):
        cell = sheet.cell(row=4, column=col_num, value=header)
        cell.font = header_font
        cell.alignment = header_alignment
        cell.fill = header_fill
        cell.border = header_border
        sheet.column_dimensions[get_column_letter(col_num)].width = 22
    sheet.row_dimensions[4].height = 30

    # Datos
    for row_num, row_data in enumerate(dataframe_to_rows(df, index=False, header=False), start=5):
        for col_num, value in enumerate(row_data, start=1):
            cell = sheet.cell(row=row_num, column=col_num, value=value)
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = data_border
            if col_num == 4:
                cell.number_format = '$#,##0.00'
        sheet.row_dimensions[row_num].height = 22
        if row_num % 2 == 0:
            for col in range(1, len(headers) + 1):
                sheet.cell(row=row_num, column=col).fill = alterno_fill

    # Ajustar Ancho de Columnas
    for col_idx, column_cells in enumerate(sheet.columns, start=1):
        max_length = 0
        column_letter = get_column_letter(col_idx)
        for cell in column_cells:
            if not isinstance(cell, MergedCell):
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except TypeError:
                    pass
        adjusted_width = (max_length + 6)
        sheet.column_dimensions[column_letter].width = adjusted_width

    # Borde alrededor de la tabla de datos
    max_row = sheet.max_row
    max_col = sheet.max_column
    borde_tabla = Border(left=Side(style='medium', color=color_borde),
                           right=Side(style='medium', color=color_borde),
                           top=Side(style='medium', color=color_borde),
                           bottom=Side(style='medium', color=color_borde))
    for row in range(4, max_row + 1):
        for col in range(1, max_col + 1):
            sheet.cell(row=row, column=col).border = borde_tabla

    # Congelar Encabezados
    sheet.freeze_panes = 'A5'

    workbook.save(filename)
    print(f"El archivo Excel '{filename}' con diseño elaborado se ha generado.")

# Generar el Excel con diseño elaborado
generar_excel_elaborado(data, r"C:\Users\PLAYER666\Downloads\agenda_cosmetologia_elaborada.xlsx")